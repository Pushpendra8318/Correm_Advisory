"""
3-Sheet Excel Generator using openpyxl.
Sheet 1: Account Details
Sheet 2: Transaction Ledger (color-coded)
Sheet 3: Category Summary & Analytics
"""

from __future__ import annotations

import re
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.parser import AccountInfo, Transaction

# ── Color palette ──────────────────────────────────────────────────────────

DARK_BLUE = "1E3A5F"
MID_BLUE = "2563EB"
LIGHT_BLUE = "DBEAFE"
GREEN_FILL = "D1FAE5"
RED_FILL = "FEE2E2"
YELLOW_FILL = "FEF9C3"
GREY_FILL = "F3F4F6"
WHITE = "FFFFFF"
BLACK = "111827"

CATEGORY_COLORS: dict[str, str] = {
    "Salary": "DCFCE7",
    "EMI / Loan": "FEE2E2",
    "Food & Dining": "FEF3C7",
    "Travel": "E0E7FF",
    "Shopping": "FCE7F3",
    "Utilities": "E5E7EB",
    "Telecom": "CFFAFE",
    "Entertainment": "EDE9FE",
    "Healthcare": "ECFDF5",
    "Education": "FFF7ED",
    "Investments": "F0FDF4",
    "Insurance": "F5F3FF",
    "Cash Withdrawal": "FEF9C3",
    "UPI / Transfer": "EFF6FF",
    "Rent": "FDF4FF",
    "Other": "F9FAFB",
}


def _make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _make_border(style="thin") -> Border:
    side = Side(style=style)
    return Border(left=side, right=side, top=side, bottom=side)


def _header_cell(ws, row: int, col: int, value: str, bold=True, bg=DARK_BLUE, fg=WHITE):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=fg, size=10)
    cell.fill = _make_fill(bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _make_border()
    return cell


def _data_cell(ws, row: int, col: int, value, bold=False, bg=None, number_format=None, align="left"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=BLACK, size=10)
    if bg:
        cell.fill = _make_fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = _make_border()
    if number_format:
        cell.number_format = number_format
    return cell


def _auto_fit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def _parse_date(date_str: str) -> Optional[datetime]:
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%d-%m-%Y", "%d-%m-%y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


def _month_label(date_str: str) -> str:
    dt = _parse_date(date_str)
    return dt.strftime("%b %Y") if dt else "Unknown"


# ── Sheet 1 — Account Details ──────────────────────────────────────────────

def _build_sheet1(wb: Workbook, account: AccountInfo) -> None:
    ws = wb.create_sheet("Account Details")
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:C1")
    title_cell = ws.cell(row=1, column=1, value="HDFC Bank — Account Summary")
    title_cell.font = Font(bold=True, color=WHITE, size=14)
    title_cell.fill = _make_fill(DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    fields = [
        ("Account Holder Name", account.holder_name or "—"),
        ("Account Number", account.account_number or "—"),
        ("Bank Name", account.bank_name),
        ("Branch", account.branch or "—"),
        ("IFSC Code", account.ifsc or "—"),
        ("Statement From", account.statement_from or "—"),
        ("Statement To", account.statement_to or "—"),
        ("Opening Balance (₹)", account.opening_balance),
        ("Closing Balance (₹)", account.closing_balance),
        ("Total Credits — Count", account.total_credit_count),
        ("Total Credits — Amount (₹)", account.total_credit_amount),
        ("Total Debits — Count", account.total_debit_count),
        ("Total Debits — Amount (₹)", account.total_debit_amount),
        ("Net Flow (₹)", round(account.total_credit_amount - account.total_debit_amount, 2)),
    ]

    for i, (label, value) in enumerate(fields, start=2):
        bg = LIGHT_BLUE if i % 2 == 0 else WHITE
        _data_cell(ws, i, 1, label, bold=True, bg=bg)
        fmt = "#,##0.00" if isinstance(value, float) else None
        _data_cell(ws, i, 2, value, bg=bg, number_format=fmt, align="right")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 4
    ws.row_dimensions[1].height = 36


# ── Sheet 2 — Transaction Ledger ──────────────────────────────────────────

LEDGER_HEADERS = [
    "Date", "Value Date", "Description", "Cheque / Ref No",
    "Deposits (CR) ₹", "Withdrawals (DR) ₹", "Running Balance ₹", "Category"
]


def _build_sheet2(wb: Workbook, transactions: list[Transaction]) -> None:
    ws = wb.create_sheet("Transaction Ledger")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    # Title
    ws.merge_cells(f"A1:{get_column_letter(len(LEDGER_HEADERS))}1")
    t = ws.cell(row=1, column=1, value="Transaction Ledger — All Transactions")
    t.font = Font(bold=True, color=WHITE, size=13)
    t.fill = _make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(LEDGER_HEADERS, start=1):
        _header_cell(ws, 2, col, header, bg=MID_BLUE)

    for row_idx, txn in enumerate(transactions, start=3):
        if txn.credit:
            row_bg = GREEN_FILL
        elif txn.debit:
            row_bg = RED_FILL
        else:
            row_bg = WHITE

        cat_bg = CATEGORY_COLORS.get(txn.category, WHITE)

        _data_cell(ws, row_idx, 1, txn.date, bg=row_bg)
        _data_cell(ws, row_idx, 2, txn.value_date, bg=row_bg)
        _data_cell(ws, row_idx, 3, txn.description, bg=row_bg)
        _data_cell(ws, row_idx, 4, txn.cheque_ref, bg=row_bg)
        _data_cell(ws, row_idx, 5, txn.credit or "", bg=row_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, row_idx, 6, txn.debit or "", bg=row_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, row_idx, 7, txn.balance, bg=row_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, row_idx, 8, txn.category, bg=cat_bg, align="center")

    # Summary totals row
    last = len(transactions) + 2
    total_row = last + 1
    ws.cell(row=total_row, column=3, value="TOTALS").font = Font(bold=True)
    cr_total = sum(t.credit for t in transactions if t.credit)
    dr_total = sum(t.debit for t in transactions if t.debit)
    c5 = ws.cell(row=total_row, column=5, value=cr_total)
    c5.font = Font(bold=True, color="15803D")
    c5.number_format = "#,##0.00"
    c6 = ws.cell(row=total_row, column=6, value=dr_total)
    c6.font = Font(bold=True, color="B91C1C")
    c6.number_format = "#,##0.00"

    _auto_fit(ws)


# ── Sheet 3 — Category Summary & Analytics ────────────────────────────────

def _section_header(ws, row: int, title: str, ncols: int = 5) -> None:
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(bold=True, color=WHITE, size=11)
    cell.fill = _make_fill(MID_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 24


def _detect_salary(transactions: list[Transaction]) -> list[Transaction]:
    """Find recurring monthly credits of similar amounts (salary pattern)."""
    from collections import defaultdict
    monthly_credits: dict[str, list[float]] = defaultdict(list)
    monthly_txns: dict[str, list[Transaction]] = defaultdict(list)

    for txn in transactions:
        if txn.credit and txn.credit > 5000:
            month = _month_label(txn.date)
            monthly_credits[month].append(txn.credit)
            monthly_txns[month].append(txn)

    # Find the most consistent large credit per month
    candidates = []
    for month, amounts in monthly_credits.items():
        for txn in monthly_txns[month]:
            if txn.credit == max(amounts):
                candidates.append(txn)

    # Check if amounts are within ±15% of each other across months
    if len(candidates) < 2:
        return []
    avg = sum(t.credit for t in candidates) / len(candidates)
    recurring = [t for t in candidates if abs(t.credit - avg) / avg <= 0.15]
    return recurring if len(recurring) >= 2 else []


def _detect_emi(transactions: list[Transaction]) -> list[dict]:
    """Detect recurring fixed monthly debits (EMI pattern)."""
    from collections import Counter
    monthly_debits: dict[str, list[Transaction]] = defaultdict(list)

    for txn in transactions:
        if txn.debit and txn.debit > 1000:
            month = _month_label(txn.date)
            monthly_debits[month].append(txn)

    # Group debits by rounded amount (±2%)
    amount_buckets: dict[int, list[Transaction]] = defaultdict(list)
    for txns in monthly_debits.values():
        for txn in txns:
            bucket = round(txn.debit / 100) * 100  # round to nearest 100
            amount_buckets[bucket].append(txn)

    emis = []
    for bucket, txns in amount_buckets.items():
        months_seen = set(_month_label(t.date) for t in txns)
        if len(months_seen) >= 2:
            avg_amount = sum(t.debit for t in txns) / len(txns)
            emis.append({
                "approximate_amount": round(avg_amount, 2),
                "occurrences": len(months_seen),
                "months": sorted(months_seen),
                "description": txns[0].description[:60],
            })

    emis.sort(key=lambda x: x["occurrences"], reverse=True)
    return emis[:5]


def _build_sheet3(wb: Workbook, transactions: list[Transaction], account: AccountInfo) -> None:
    from app.categorizer import get_categorization_stats

    ws = wb.create_sheet("Analytics")
    ws.sheet_view.showGridLines = False

    current_row = 1

    # ── Title ──────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1, value="Statement Analytics & Category Summary")
    t.font = Font(bold=True, color=WHITE, size=14)
    t.fill = _make_fill(DARK_BLUE)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34
    current_row = 2

    # ── Section 1: Category Summary ────────────────────────────────────────
    current_row += 1
    _section_header(ws, current_row, "1.  Category-Wise Summary", 5)
    current_row += 1

    cat_headers = ["Category", "Total Credits (₹)", "Total Debits (₹)", "Net (₹)", "Transaction Count"]
    for col, h in enumerate(cat_headers, start=1):
        _header_cell(ws, current_row, col, h, bg="374151")
    current_row += 1

    cat_data: dict[str, dict] = defaultdict(lambda: {"credit": 0.0, "debit": 0.0, "count": 0})
    for txn in transactions:
        cat_data[txn.category]["count"] += 1
        if txn.credit:
            cat_data[txn.category]["credit"] += txn.credit
        if txn.debit:
            cat_data[txn.category]["debit"] += txn.debit

    for i, (cat, data) in enumerate(sorted(cat_data.items())):
        bg = CATEGORY_COLORS.get(cat, WHITE)
        alt_bg = GREY_FILL if i % 2 else WHITE
        _data_cell(ws, current_row, 1, cat, bold=True, bg=bg)
        _data_cell(ws, current_row, 2, round(data["credit"], 2), bg=alt_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, current_row, 3, round(data["debit"], 2), bg=alt_bg, number_format="#,##0.00", align="right")
        net = round(data["credit"] - data["debit"], 2)
        net_bg = GREEN_FILL if net >= 0 else RED_FILL
        _data_cell(ws, current_row, 4, net, bg=net_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, current_row, 5, data["count"], bg=alt_bg, align="center")
        current_row += 1

    current_row += 1

    # ── Section 2: Month-wise Breakdown ────────────────────────────────────
    _section_header(ws, current_row, "2.  Month-Wise Inflow / Outflow", 5)
    current_row += 1

    month_headers = ["Month", "Total Inflow (₹)", "Total Outflow (₹)", "Net (₹)", "Transaction Count"]
    for col, h in enumerate(month_headers, start=1):
        _header_cell(ws, current_row, col, h, bg="374151")
    current_row += 1

    monthly: dict[str, dict] = defaultdict(lambda: {"inflow": 0.0, "outflow": 0.0, "count": 0})
    for txn in transactions:
        month = _month_label(txn.date)
        monthly[month]["count"] += 1
        if txn.credit:
            monthly[month]["inflow"] += txn.credit
        if txn.debit:
            monthly[month]["outflow"] += txn.debit

    for i, (month, data) in enumerate(sorted(monthly.items(), key=lambda x: _parse_date("01/" + x[0].replace(" ", "/")) or datetime.min)):
        alt_bg = LIGHT_BLUE if i % 2 else WHITE
        net = round(data["inflow"] - data["outflow"], 2)
        _data_cell(ws, current_row, 1, month, bold=True, bg=alt_bg)
        _data_cell(ws, current_row, 2, round(data["inflow"], 2), bg=alt_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, current_row, 3, round(data["outflow"], 2), bg=alt_bg, number_format="#,##0.00", align="right")
        net_bg = GREEN_FILL if net >= 0 else RED_FILL
        _data_cell(ws, current_row, 4, net, bg=net_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, current_row, 5, data["count"], bg=alt_bg, align="center")
        current_row += 1

    current_row += 1

    # ── Section 3: Top 5 Transactions ──────────────────────────────────────
    _section_header(ws, current_row, "3.  Top 5 Largest Transactions", 5)
    current_row += 1

    top5_headers = ["#", "Date", "Description", "Amount (₹)", "Type"]
    for col, h in enumerate(top5_headers, start=1):
        _header_cell(ws, current_row, col, h, bg="374151")
    current_row += 1

    def _abs_amount(t: Transaction) -> float:
        return max(t.credit or 0, t.debit or 0)

    top5 = sorted(transactions, key=_abs_amount, reverse=True)[:5]
    for i, txn in enumerate(top5, start=1):
        amount = txn.credit if txn.credit else txn.debit
        txn_type = "Credit" if txn.credit else "Debit"
        row_bg = GREEN_FILL if txn.credit else RED_FILL
        _data_cell(ws, current_row, 1, i, bg=row_bg, align="center")
        _data_cell(ws, current_row, 2, txn.date, bg=row_bg)
        _data_cell(ws, current_row, 3, txn.description[:70], bg=row_bg)
        _data_cell(ws, current_row, 4, amount, bg=row_bg, number_format="#,##0.00", align="right")
        _data_cell(ws, current_row, 5, txn_type, bg=row_bg, align="center")
        current_row += 1

    current_row += 1

    # ── Section 4: Salary Detection ────────────────────────────────────────
    _section_header(ws, current_row, "4.  Salary Detection (Recurring Monthly Credits)", 5)
    current_row += 1

    sal_txns = _detect_salary(transactions)
    if sal_txns:
        for col, h in enumerate(["Date", "Description", "Amount (₹)", "Month", "Notes"], start=1):
            _header_cell(ws, current_row, col, h, bg="374151")
        current_row += 1
        for txn in sal_txns:
            _data_cell(ws, current_row, 1, txn.date, bg=GREEN_FILL)
            _data_cell(ws, current_row, 2, txn.description[:60], bg=GREEN_FILL)
            _data_cell(ws, current_row, 3, txn.credit, bg=GREEN_FILL, number_format="#,##0.00", align="right")
            _data_cell(ws, current_row, 4, _month_label(txn.date), bg=GREEN_FILL)
            _data_cell(ws, current_row, 5, "Auto-detected salary", bg=GREEN_FILL)
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="No recurring salary detected in this statement period.")
        current_row += 1

    current_row += 1

    # ── Section 5: EMI Detection ───────────────────────────────────────────
    _section_header(ws, current_row, "5.  EMI / Loan Detection (Recurring Fixed Debits)", 5)
    current_row += 1

    emi_list = _detect_emi(transactions)
    if emi_list:
        for col, h in enumerate(["Approx Amount (₹)", "Description", "Occurrences", "Months Detected", ""], start=1):
            _header_cell(ws, current_row, col, h, bg="374151")
        current_row += 1
        for emi in emi_list:
            _data_cell(ws, current_row, 1, emi["approximate_amount"], bg=RED_FILL, number_format="#,##0.00", align="right")
            _data_cell(ws, current_row, 2, emi["description"], bg=RED_FILL)
            _data_cell(ws, current_row, 3, emi["occurrences"], bg=RED_FILL, align="center")
            _data_cell(ws, current_row, 4, ", ".join(emi["months"]), bg=RED_FILL)
            _data_cell(ws, current_row, 5, "Auto-detected EMI/Recurring debit", bg=RED_FILL)
            current_row += 1
    else:
        ws.cell(row=current_row, column=1, value="No recurring EMI pattern detected.")
        current_row += 1

    current_row += 1

    # ── Section 6: Categorization % ────────────────────────────────────────
    _section_header(ws, current_row, "6.  Categorization Coverage", 5)
    current_row += 1

    stats = get_categorization_stats(transactions)
    pct_rows = [
        ("Total Transactions", stats["total"]),
        ("Categorized", stats["categorized"]),
        ("Uncategorized (Other)", stats["uncategorized"]),
        ("Categorization %", f"{stats['pct_categorized']}%"),
    ]
    for label, value in pct_rows:
        _data_cell(ws, current_row, 1, label, bold=True, bg=LIGHT_BLUE)
        _data_cell(ws, current_row, 2, value, bg=WHITE, align="center")
        current_row += 1

    _auto_fit(ws)


# ── Public entry point ────────────────────────────────────────────────────

def generate_excel(account: AccountInfo, transactions: list[Transaction]) -> bytes:
    """Build and return the 3-sheet Excel file as bytes."""
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    _build_sheet1(wb, account)
    _build_sheet2(wb, transactions)
    _build_sheet3(wb, transactions, account)

    # Set tab colors
    wb["Account Details"].sheet_properties.tabColor = "1E3A5F"
    wb["Transaction Ledger"].sheet_properties.tabColor = "2563EB"
    wb["Analytics"].sheet_properties.tabColor = "059669"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
